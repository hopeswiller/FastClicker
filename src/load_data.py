"""Implements all methods to load data from excel file."""
__all__ = ["get_click_profile", "get_nth_key"]

import os
import sys
import openpyxl as xl


def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.environ.get("_MEIPASS2", os.path.abspath("."))

    return os.path.join(base_path, relative_path)


def get_nth_key(dictionary, n=0):
    if n < 0:
        n += len(dictionary)
    for i, key in enumerate(dictionary.keys()):
        if i == n:
            return key
    raise IndexError("dictionary index out of range")


def get_click_profile(path):
    path = resource_path(path)
    book = xl.load_workbook(path)
    # get active sheet
    sheet = book.active
    profile = []

    headers = {}
    for col in sheet.iter_cols(max_col=sheet.max_column):
        headers[str(col[0].value).strip()] = None

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data = {}
        if all(row):
            for index, col in enumerate(row):
                name = get_nth_key(headers, index)
                data[str(name).strip()] = col

            # add row data to profile
            profile.append(data)

    return profile, headers
