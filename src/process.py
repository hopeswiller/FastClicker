import openpyxl as xl

file_path = "profile.xlsx"

book = xl.load_workbook(file_path)
# get active sheet
sheet = book.active
profile = []


def get_nth_key(dictionary, n=0):
    if n < 0:
        n += len(dictionary)
    for i, key in enumerate(dictionary.keys()):
        if i == n:
            return key
    raise IndexError("dictionary index out of range")


def get_click_profile():
    headers = {}
    for col in sheet.iter_cols(max_col=sheet.max_column):
        headers[col[0].value] = None

    for row in sheet.iter_rows(min_row=2):
        data = {}
        for index, col in enumerate(row):
            name = get_nth_key(headers, index)
            data[str(name).strip()] = col.value

        # add row data to profile
        profile.append(data)

    del headers
    return profile
